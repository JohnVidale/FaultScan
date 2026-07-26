"""Safely apply measured event-stack residuals to catalog time shifts."""

from __future__ import annotations

import argparse
import math
import os
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


DEFAULT_CATALOG = Path(
    "/Users/jvidale/Documents/Research/FaultScanR/event_sta_info/catalog_local_hand.xlsx"
)
DEFAULT_OUTPUT_DIR = Path(
    "/Users/jvidale/Documents/Research/FaultScanR/output/20260718_174326_0912"
)
DEFAULT_REFERENCE_EVENT = "CI_40353472"


def _as_measured(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not pd.isna(value):
        return bool(value)
    return str(value).strip().lower() in {"true", "yes", "1"}


def _finite_float(value: object, label: str, event_id: str) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{event_id}: {label} is not numeric: {value!r}") from exc
    if not math.isfinite(number):
        raise ValueError(f"{event_id}: {label} is not finite: {value!r}")
    return number


def compute_catalog_updates(
    catalog: pd.DataFrame,
    shifts: pd.DataFrame,
    destination_column: str,
    *,
    source_column: str | None = None,
    overwrite_derived: bool = False,
    tolerance: float = 1e-9,
) -> tuple[list[dict[str, object]], str]:
    """Validate provenance and return idempotent catalog updates.

    The alignment shift is a residual measured after applying the catalog value
    recorded in the workbook. Therefore the proposed absolute catalog value is
    ``catalog_time_shift_seconds + shift_left_to_align_waveform_seconds``.
    """
    required_catalog = {"evid"}
    required_shifts = {
        "event_id",
        "shift_left_to_align_waveform_seconds",
        "xcorr_residual_measured",
        "catalog_time_shift_seconds",
    }
    missing_catalog = required_catalog - set(catalog.columns)
    missing_shifts = required_shifts - set(shifts.columns)
    if missing_catalog:
        raise ValueError(f"Catalog is missing columns: {sorted(missing_catalog)}")
    if missing_shifts:
        raise ValueError(f"Alignment workbook is missing columns: {sorted(missing_shifts)}")

    shifts = shifts.copy()
    shifts["event_id"] = shifts["event_id"].astype(str).str.strip()
    if shifts["event_id"].eq("").any() or shifts["event_id"].duplicated().any():
        raise ValueError("Alignment workbook contains blank or duplicate event IDs")

    unmeasured = shifts.loc[
        ~shifts["xcorr_residual_measured"].map(_as_measured), "event_id"
    ].tolist()
    if unmeasured:
        raise ValueError(
            "Refusing to update from a workbook without measured cross-correlation "
            f"residuals; affected events: {', '.join(unmeasured)}"
        )

    if "catalog_time_shift_column" in shifts.columns:
        recorded_columns = {
            str(value).strip()
            for value in shifts["catalog_time_shift_column"].dropna()
            if str(value).strip()
        }
        if len(recorded_columns) != 1:
            raise ValueError(
                "Alignment workbook must record exactly one catalog_time_shift_column"
            )
        recorded_source = recorded_columns.pop()
        if source_column is not None and source_column != recorded_source:
            raise ValueError(
                f"Requested source column {source_column!r} does not match workbook "
                f"provenance {recorded_source!r}"
            )
        source_column = recorded_source
    elif source_column is None:
        raise ValueError(
            "Alignment workbook lacks catalog_time_shift_column provenance; "
            "supply --source-column only after verifying the older workbook manually"
        )

    catalog = catalog.copy()
    catalog["evid"] = catalog["evid"].astype(str).str.strip()
    if catalog["evid"].duplicated().any():
        raise ValueError("Catalog contains duplicate event IDs")
    catalog_by_event = catalog.set_index("evid", drop=False)

    updates: list[dict[str, object]] = []
    missing_events: list[str] = []
    for row in shifts.to_dict("records"):
        event_id = str(row["event_id"]).strip()
        if event_id not in catalog_by_event.index:
            missing_events.append(event_id)
            continue

        baseline = _finite_float(
            row["catalog_time_shift_seconds"], "catalog baseline", event_id
        )
        residual = _finite_float(
            row["shift_left_to_align_waveform_seconds"], "alignment residual", event_id
        )
        proposed = baseline + residual
        catalog_row = catalog_by_event.loc[event_id]
        current_value = catalog_row.get(destination_column, float("nan"))
        current = (
            float(current_value)
            if not pd.isna(current_value) and math.isfinite(float(current_value))
            else None
        )

        if current is not None and math.isclose(current, proposed, abs_tol=tolerance):
            action = "unchanged"
        elif destination_column == source_column:
            if current is None or not math.isclose(current, baseline, abs_tol=tolerance):
                raise ValueError(
                    f"{event_id}: catalog {destination_column!r} is {current!r}, but the "
                    f"workbook used {baseline:.9g}; regenerate alignment results before updating"
                )
            action = "update"
        elif current is None or math.isclose(current, baseline, abs_tol=tolerance):
            action = "update"
        elif overwrite_derived:
            action = "update"
        else:
            raise ValueError(
                f"{event_id}: derived destination {destination_column!r} already contains "
                f"{current:.9g}; use --overwrite-derived only after checking its provenance"
            )

        updates.append(
            {
                "event_id": event_id,
                "baseline_seconds": baseline,
                "residual_seconds": residual,
                "proposed_seconds": proposed,
                "action": action,
            }
        )

    if missing_events:
        raise ValueError(
            "Alignment events missing from catalog: " + ", ".join(sorted(missing_events))
        )
    return updates, source_column


def apply_updates_to_workbook(
    catalog_path: Path,
    destination_column: str,
    updates: list[dict[str, object]],
) -> int:
    """Apply validated updates while preserving the existing workbook layout."""
    workbook = load_workbook(catalog_path)
    sheet = workbook[workbook.sheetnames[0]]
    headers = {
        str(cell.value).strip(): column
        for column, cell in enumerate(sheet[1], start=1)
        if cell.value is not None
    }
    if "evid" not in headers:
        raise ValueError('Catalog worksheet is missing the "evid" column')
    if destination_column not in headers:
        destination_index = sheet.max_column + 1
        sheet.cell(1, destination_index, destination_column)
    else:
        destination_index = headers[destination_column]

    update_by_event = {
        str(row["event_id"]): row
        for row in updates
        if row["action"] == "update"
    }
    updated = 0
    for row_index in range(2, sheet.max_row + 1):
        event_id = str(sheet.cell(row_index, headers["evid"]).value or "").strip()
        if event_id not in update_by_event:
            continue
        cell = sheet.cell(row_index, destination_index)
        cell.value = float(update_by_event[event_id]["proposed_seconds"])
        cell.number_format = "0.000000"
        updated += 1

    fd, temporary_name = tempfile.mkstemp(
        prefix=f".{catalog_path.stem}_", suffix=catalog_path.suffix, dir=catalog_path.parent
    )
    os.close(fd)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, catalog_path)
    finally:
        temporary_path.unlink(missing_ok=True)
    return updated


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Add measured event-stack residuals to the catalog shifts used by that run. "
            "Existing or stale values are rejected by default."
        )
    )
    parser.add_argument("component", nargs="?", default="R", choices=("R", "T", "Z"))
    parser.add_argument("output_dir", nargs="?", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--catalog", type=Path, default=DEFAULT_CATALOG)
    parser.add_argument("--reference-event", default=DEFAULT_REFERENCE_EVENT)
    parser.add_argument("--source-column")
    parser.add_argument("--destination-column")
    parser.add_argument("--overwrite-derived", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    component = args.component.upper()
    shifts_path = args.output_dir / (
        f"all_events_{component}_stack_xcorr_alignment_to_"
        f"{args.reference_event}.xlsx"
    )
    destination = args.destination_column or (
        "time shift" if component == "R" else f"time shift {component}"
    )
    catalog = pd.read_excel(args.catalog)
    shifts = pd.read_excel(shifts_path)
    updates, source = compute_catalog_updates(
        catalog,
        shifts,
        destination,
        source_column=args.source_column,
        overwrite_derived=args.overwrite_derived,
    )
    changed = sum(row["action"] == "update" for row in updates)

    print(f"Alignment workbook: {shifts_path}")
    print(f"Catalog shift used by run: {source}")
    print(f"Destination catalog column: {destination}")
    print(f"Rows requiring update: {changed}; already current: {len(updates) - changed}")
    for row in updates:
        if row["action"] == "update":
            print(
                f"  {row['event_id']}: {row['baseline_seconds']:+.6f} + "
                f"{row['residual_seconds']:+.6f} = {row['proposed_seconds']:+.6f}"
            )

    if args.dry_run:
        print("Dry run only; catalog was not changed.")
        return
    updated = apply_updates_to_workbook(args.catalog, destination, updates)
    print(f"Updated {updated} rows in {args.catalog}")


if __name__ == "__main__":
    main()
